from __future__ import annotations

from datetime import date
from importlib import resources
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from .config import WORKBOOK_HEADERS
from .sync import SyncResult, sync_netzanmeldungen


def run_demo(output_dir: str | Path = "demo_output") -> SyncResult:
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    workbook = output_path / "netzanmeldungen_demo.xlsx"
    result_workbook = output_path / "netzanmeldungen_aktualisiert.xlsx"
    create_demo_workbook(workbook)

    sample_root = resources.files("netzanmeldung_excel_sync.sample_data")
    return sync_netzanmeldungen(
        termine_csv=sample_root / "termine.csv",
        projekte_csv=sample_root / "projekte.csv",
        kontakte_csv=sample_root / "kontakte.csv",
        workbook_path=workbook,
        output_path=result_workbook,
    )


def create_demo_workbook(path: str | Path) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Netzanmeldungen"
    ws.append(list(WORKBOOK_HEADERS))
    ws.append(
        [
            "PV-DEMO-1001",
            "Demo",
            "Kontakt-1001",
            date(2025, 4, 8),
            "Demo-Netz",
            "gestellt",
            None,
            None,
            None,
            None,
            None,
            None,
            "bereits vorhanden",
            None,
        ]
    )
    ws.append(
        [
            "PV-DEMO-1004",
            "Demo",
            "Kontakt-1004",
            date(2025, 4, 10),
            "Demo-Netz",
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            "leere Dublette",
            None,
        ]
    )
    ws.append(
        [
            "PV-DEMO-1004",
            "Demo",
            "Kontakt-1004",
            date(2025, 4, 10),
            "Demo-Netz",
            "gestellt",
            None,
            None,
            None,
            None,
            None,
            None,
            "gefuellte Dublette",
            None,
        ]
    )

    for row in range(2, ws.max_row + 1):
        ws.cell(row, 4).number_format = "yyyy-mm-dd"

    table = Table(displayName="Tabelle1", ref=f"A1:N{ws.max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    wb.save(path)
    return path
