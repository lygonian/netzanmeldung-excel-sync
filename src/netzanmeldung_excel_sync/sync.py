from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Sequence

import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

from .config import (
    DEFAULT_MIN_START_DATE,
    DEFAULT_MOUNT_TYPES,
    DEFAULT_TABLE_NAME,
    KONTAKTE_REQUIRED_COLUMNS,
    OUTPUT_COLUMNS,
    PROJEKTE_REQUIRED_COLUMNS,
    TERMINE_REQUIRED_COLUMNS,
)


@dataclass(frozen=True)
class SyncResult:
    eligible_projects: int
    added_rows: int
    deleted_rows: int
    table_ref: str
    output_path: Path


def prepare_new_rows(
    termine_csv: str | Path,
    projekte_csv: str | Path,
    kontakte_csv: str | Path,
    *,
    min_start_date: str = DEFAULT_MIN_START_DATE,
    mount_types: Sequence[str] = DEFAULT_MOUNT_TYPES,
    encoding: str = "utf-8-sig",
) -> pd.DataFrame:
    """Build the project rows that are eligible for insertion into Excel."""

    termine = pd.read_csv(termine_csv, encoding=encoding)
    projekte = pd.read_csv(projekte_csv, encoding=encoding)
    kontakte = pd.read_csv(kontakte_csv, encoding=encoding)

    _require_columns(termine, TERMINE_REQUIRED_COLUMNS, "termine_csv")
    _require_columns(projekte, PROJEKTE_REQUIRED_COLUMNS, "projekte_csv")
    _require_columns(kontakte, KONTAKTE_REQUIRED_COLUMNS, "kontakte_csv")

    termine = termine.copy()
    termine["start"] = _parse_start_dates(termine["start"])
    termine["project_match_id"] = pd.to_numeric(termine["project_match_id"], errors="coerce").astype("Int64")

    eligible = termine[
        termine["localized_type"].isin(mount_types)
        & (~_normalize_deleted(termine["deleted"]))
        & (termine["start"] >= pd.Timestamp(min_start_date))
        & termine["project_match_id"].notna()
    ].copy()

    projekte = projekte[["contact_id", "project_nr", "id"]].copy()
    projekte["id"] = pd.to_numeric(projekte["id"], errors="coerce").astype("Int64")
    projekte["contact_id"] = pd.to_numeric(projekte["contact_id"], errors="coerce").astype("Int64")

    kontakte = kontakte[["first_name", "last_name", "id"]].copy()
    kontakte["id"] = pd.to_numeric(kontakte["id"], errors="coerce").astype("Int64")

    merged = pd.merge(
        eligible[["project_match_id", "start"]],
        projekte,
        left_on="project_match_id",
        right_on="id",
        how="left",
    )
    merged = merged.drop_duplicates(subset=["project_nr"], keep="first")
    merged = pd.merge(
        merged,
        kontakte,
        left_on="contact_id",
        right_on="id",
        how="left",
        suffixes=("_project", "_contact"),
    )
    merged["start_datum"] = merged["start"].dt.date

    rows = merged[list(OUTPUT_COLUMNS)].dropna(subset=["project_nr"]).reset_index(drop=True)
    return rows


def update_workbook(
    workbook_path: str | Path,
    rows: pd.DataFrame,
    *,
    output_path: str | Path | None = None,
    table_name: str = DEFAULT_TABLE_NAME,
) -> SyncResult:
    """Insert missing project rows and clean duplicate empty request rows."""

    _require_columns(rows, OUTPUT_COLUMNS, "rows")
    workbook = Path(workbook_path)
    output = Path(output_path) if output_path is not None else workbook
    output.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(workbook)
    ws = wb.active
    if table_name not in ws.tables:
        available = ", ".join(ws.tables.keys()) or "keine"
        raise ValueError(f"Excel-Tabelle {table_name!r} nicht gefunden; vorhanden: {available}")

    table = ws.tables[table_name]
    min_col, min_row, max_col, _ = range_boundaries(table.ref)
    existing = _existing_project_keys(ws, min_row=min_row, min_col=min_col)

    added = 0
    for record in rows.to_dict("records"):
        project_key = _project_key(record["project_nr"])
        if not project_key or project_key in existing:
            continue

        target_row = ws.max_row + 1
        if target_row > min_row + 1:
            _copy_row_format(ws, source_row=target_row - 1, target_row=target_row, min_col=min_col, max_col=max_col)

        ws.cell(target_row, min_col).value = record["project_nr"]
        ws.cell(target_row, min_col + 1).value = record["first_name"]
        ws.cell(target_row, min_col + 2).value = record["last_name"]
        ws.cell(target_row, min_col + 3).value = record["start_datum"]
        existing.add(project_key)
        added += 1

    deleted = _delete_empty_request_duplicates(ws, min_row=min_row, project_col=min_col, request_col=min_col + 5)

    new_last_row = max(ws.max_row, min_row)
    table.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{new_last_row}"
    wb.save(output)

    return SyncResult(
        eligible_projects=len(rows),
        added_rows=added,
        deleted_rows=deleted,
        table_ref=table.ref,
        output_path=output,
    )


def sync_netzanmeldungen(
    *,
    termine_csv: str | Path,
    projekte_csv: str | Path,
    kontakte_csv: str | Path,
    workbook_path: str | Path,
    output_path: str | Path | None = None,
    table_name: str = DEFAULT_TABLE_NAME,
    min_start_date: str = DEFAULT_MIN_START_DATE,
) -> SyncResult:
    rows = prepare_new_rows(
        termine_csv,
        projekte_csv,
        kontakte_csv,
        min_start_date=min_start_date,
    )
    return update_workbook(workbook_path, rows, output_path=output_path, table_name=table_name)


def _require_columns(frame: pd.DataFrame, required: Iterable[str], source_name: str) -> None:
    missing = [column for column in required if column not in frame.columns]
    if missing:
        joined = ", ".join(missing)
        raise ValueError(f"{source_name} fehlt Pflichtspalte(n): {joined}")


def _normalize_deleted(series: pd.Series) -> pd.Series:
    if pd.api.types.is_bool_dtype(series):
        return series.fillna(False)

    truthy = {"1", "true", "yes", "ja", "y", "j", "x"}
    normalized = series.astype("string").str.strip().str.lower()
    return normalized.isin(truthy).fillna(False)


def _parse_start_dates(series: pd.Series) -> pd.Series:
    def parse_one(value: object) -> pd.Timestamp:
        if value is None or pd.isna(value):
            return pd.NaT
        try:
            timestamp = pd.Timestamp(value)
        except (TypeError, ValueError):
            return pd.NaT
        if timestamp.tzinfo is not None:
            return timestamp.tz_localize(None)
        return timestamp

    return pd.to_datetime(series.map(parse_one), errors="coerce")


def _existing_project_keys(ws, *, min_row: int, min_col: int) -> set[str]:
    keys: set[str] = set()
    for row_number in range(min_row + 1, ws.max_row + 1):
        key = _project_key(ws.cell(row_number, min_col).value)
        if key:
            keys.add(key)
    return keys


def _delete_empty_request_duplicates(ws, *, min_row: int, project_col: int, request_col: int) -> int:
    rows_by_project: dict[str, list[tuple[int, object]]] = {}
    for row_number in range(min_row + 1, ws.max_row + 1):
        project = _project_key(ws.cell(row_number, project_col).value)
        if project:
            rows_by_project.setdefault(project, []).append((row_number, ws.cell(row_number, request_col).value))

    rows_to_delete: list[int] = []
    for entries in rows_by_project.values():
        has_request = any(not _is_blank(request_value) for _, request_value in entries)
        if has_request:
            rows_to_delete.extend(row_number for row_number, request_value in entries if _is_blank(request_value))

    for row_number in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row_number)

    return len(rows_to_delete)


def _copy_row_format(ws, *, source_row: int, target_row: int, min_col: int, max_col: int) -> None:
    for column in range(min_col, max_col + 1):
        source = ws.cell(source_row, column)
        target = ws.cell(target_row, column)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.protection:
            target.protection = copy(source.protection)


def _project_key(value: object) -> str:
    if value is None or pd.isna(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _is_blank(value: object) -> bool:
    return value is None or str(value).strip() == ""
