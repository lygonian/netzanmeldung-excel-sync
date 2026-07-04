from __future__ import annotations

import csv
from datetime import date

import openpyxl
import pytest
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from netzanmeldung_excel_sync.config import WORKBOOK_HEADERS
from netzanmeldung_excel_sync.sync import prepare_new_rows, sync_netzanmeldungen, update_workbook


def test_prepare_new_rows_filters_and_merges(tmp_path):
    termine = tmp_path / "termine.csv"
    projekte = tmp_path / "projekte.csv"
    kontakte = tmp_path / "kontakte.csv"

    _write_csv(
        termine,
        ["id", "project_match_id", "localized_type", "deleted", "start"],
        [
            {
                "id": "1",
                "project_match_id": "10",
                "localized_type": "DC-Montage ohne Geruest",
                "deleted": "false",
                "start": "2025-04-02T08:00:00+02:00",
            },
            {
                "id": "2",
                "project_match_id": "11",
                "localized_type": "DC-Montage ohne Geruest",
                "deleted": "true",
                "start": "2025-04-03T08:00:00+02:00",
            },
            {
                "id": "3",
                "project_match_id": "12",
                "localized_type": "AC-Montage",
                "deleted": "false",
                "start": "2025-04-04T08:00:00+02:00",
            },
            {
                "id": "4",
                "project_match_id": "13",
                "localized_type": "DC-Montage mit Geruest",
                "deleted": "false",
                "start": "2025-03-30T08:00:00+02:00",
            },
        ],
    )
    _write_csv(
        projekte,
        ["id", "contact_id", "project_nr"],
        [
            {"id": "10", "contact_id": "100", "project_nr": "PV-DEMO-001"},
            {"id": "11", "contact_id": "101", "project_nr": "PV-DEMO-002"},
            {"id": "12", "contact_id": "102", "project_nr": "PV-DEMO-003"},
            {"id": "13", "contact_id": "103", "project_nr": "PV-DEMO-004"},
        ],
    )
    _write_csv(
        kontakte,
        ["id", "first_name", "last_name"],
        [
            {"id": "100", "first_name": "Demo", "last_name": "Kontakt-001"},
            {"id": "101", "first_name": "Demo", "last_name": "Kontakt-002"},
            {"id": "102", "first_name": "Demo", "last_name": "Kontakt-003"},
            {"id": "103", "first_name": "Demo", "last_name": "Kontakt-004"},
        ],
    )

    rows = prepare_new_rows(termine, projekte, kontakte)

    assert rows["project_nr"].tolist() == ["PV-DEMO-001"]
    assert rows.loc[0, "first_name"] == "Demo"
    assert rows.loc[0, "last_name"] == "Kontakt-001"
    assert rows.loc[0, "start_datum"] == date(2025, 4, 2)


def test_prepare_new_rows_keeps_local_timezone_date_at_cutoff(tmp_path):
    termine = tmp_path / "termine.csv"
    projekte = tmp_path / "projekte.csv"
    kontakte = tmp_path / "kontakte.csv"

    _write_csv(
        termine,
        ["id", "project_match_id", "localized_type", "deleted", "start"],
        [
            {
                "id": "1",
                "project_match_id": "10",
                "localized_type": "DC-Montage ohne Geruest",
                "deleted": "false",
                "start": "2025-04-01T00:30:00+02:00",
            },
        ],
    )
    _write_csv(projekte, ["id", "contact_id", "project_nr"], [{"id": "10", "contact_id": "100", "project_nr": "PV-DEMO-001"}])
    _write_csv(kontakte, ["id", "first_name", "last_name"], [{"id": "100", "first_name": "Demo", "last_name": "Kontakt-001"}])

    rows = prepare_new_rows(termine, projekte, kontakte)

    assert rows["project_nr"].tolist() == ["PV-DEMO-001"]
    assert rows.loc[0, "start_datum"] == date(2025, 4, 1)


def test_sync_adds_missing_rows_and_removes_blank_duplicate(tmp_path):
    termine = tmp_path / "termine.csv"
    projekte = tmp_path / "projekte.csv"
    kontakte = tmp_path / "kontakte.csv"
    workbook = tmp_path / "netzanmeldungen.xlsx"
    output = tmp_path / "out.xlsx"

    _write_csv(
        termine,
        ["id", "project_match_id", "localized_type", "deleted", "start"],
        [
            {
                "id": "1",
                "project_match_id": "101",
                "localized_type": "DC-Montage ohne Geruest",
                "deleted": "false",
                "start": "2025-04-08T08:00:00+02:00",
            },
            {
                "id": "2",
                "project_match_id": "103",
                "localized_type": "DC-Montage mit Geruest",
                "deleted": "false",
                "start": "2025-04-15T09:30:00+02:00",
            },
        ],
    )
    _write_csv(
        projekte,
        ["id", "contact_id", "project_nr"],
        [
            {"id": "101", "contact_id": "201", "project_nr": "PV-DEMO-1001"},
            {"id": "103", "contact_id": "203", "project_nr": "PV-DEMO-1003"},
        ],
    )
    _write_csv(
        kontakte,
        ["id", "first_name", "last_name"],
        [
            {"id": "201", "first_name": "Demo", "last_name": "Kontakt-1001"},
            {"id": "203", "first_name": "Demo", "last_name": "Kontakt-1003"},
        ],
    )
    _create_workbook(
        workbook,
        [
            ["PV-DEMO-1001", "Demo", "Kontakt-1001", date(2025, 4, 8), "Demo-Netz", "gestellt"],
            ["PV-DEMO-1002", "Demo", "Kontakt-1002", date(2025, 4, 9), "Demo-Netz", None],
            ["PV-DEMO-1002", "Demo", "Kontakt-1002", date(2025, 4, 9), "Demo-Netz", "gestellt"],
        ],
    )

    result = sync_netzanmeldungen(
        termine_csv=termine,
        projekte_csv=projekte,
        kontakte_csv=kontakte,
        workbook_path=workbook,
        output_path=output,
    )

    assert result.eligible_projects == 2
    assert result.added_rows == 1
    assert result.deleted_rows == 1
    assert result.table_ref == "A1:N4"

    wb = openpyxl.load_workbook(output)
    ws = wb.active
    projects = [ws.cell(row, 1).value for row in range(2, ws.max_row + 1)]
    assert projects == ["PV-DEMO-1001", "PV-DEMO-1002", "PV-DEMO-1003"]
    assert ws.cell(4, 2).value == "Demo"
    assert ws.cell(4, 3).value == "Kontakt-1003"


def test_update_workbook_requires_named_table(tmp_path):
    workbook = tmp_path / "netzanmeldungen.xlsx"
    wb = Workbook()
    wb.active.append(list(WORKBOOK_HEADERS))
    wb.save(workbook)

    rows = prepare_new_rows(
        _csv_with_rows(tmp_path, "termine.csv", ["project_match_id", "start", "localized_type", "deleted"], []),
        _csv_with_rows(tmp_path, "projekte.csv", ["id", "contact_id", "project_nr"], []),
        _csv_with_rows(tmp_path, "kontakte.csv", ["id", "first_name", "last_name"], []),
    )

    with pytest.raises(ValueError, match="Tabelle1"):
        update_workbook(workbook, rows)


def _write_csv(path, fieldnames, rows):
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def _csv_with_rows(tmp_path, name, fieldnames, rows):
    path = tmp_path / name
    _write_csv(path, fieldnames, rows)
    return path


def _create_workbook(path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Netzanmeldungen"
    ws.append(list(WORKBOOK_HEADERS))
    for row in rows:
        padded = list(row) + [None] * (len(WORKBOOK_HEADERS) - len(row))
        ws.append(padded)
    for row_number in range(2, ws.max_row + 1):
        ws.cell(row_number, 4).number_format = "yyyy-mm-dd"

    table = Table(displayName="Tabelle1", ref=f"A1:N{ws.max_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    wb.save(path)
